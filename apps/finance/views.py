"""
Finance Views - Management Information System
"""

from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django_filters import rest_framework as filters
from django.db.models import Sum, Q
from django.http import HttpResponse
from datetime import datetime
import csv
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from .models import (
    Account, JournalEntry, JournalEntryItem, IncomeCategory, Income,
    ExpenseCategory, Expense, Invoice, Payment
)
from .serializers import (
    AccountSerializer, JournalEntryListSerializer, JournalEntryDetailSerializer,
    IncomeCategorySerializer, IncomeSerializer, ExpenseCategorySerializer,
    ExpenseSerializer, InvoiceSerializer, InvoiceListSerializer,
    PaymentSerializer
)


class AccountFilter(filters.FilterSet):
    account_type = filters.ChoiceFilter(choices=Account.ACCOUNT_TYPE_CHOICES)
    is_cash = filters.BooleanFilter()
    is_active = filters.BooleanFilter()
    
    class Meta:
        model = Account
        fields = ['account_type', 'is_cash', 'is_active']


class AccountViewSet(viewsets.ModelViewSet):
    """ViewSet untuk Account/COA"""
    queryset = Account.objects.filter(parent__isnull=True)
    serializer_class = AccountSerializer
    permission_classes = [IsAuthenticated]
    filterset_class = AccountFilter
    search_fields = ['code', 'name']
    ordering = ['code']
    
    @action(detail=False, methods=['get'])
    def cash_accounts(self, request):
        """Get cash/bank accounts"""
        accounts = Account.objects.filter(is_cash=True, is_active=True)
        return Response(AccountSerializer(accounts, many=True).data)
    
    @action(detail=False, methods=['get'])
    def stats(self, request):
        """Account statistics"""
        total = Account.objects.count()
        active = Account.objects.filter(is_active=True).count()
        cash = Account.objects.filter(is_cash=True).count()
        
        return Response({
            'total': total,
            'active': active,
            'cash_accounts': cash
        })


class JournalEntryViewSet(viewsets.ModelViewSet):
    """ViewSet untuk Journal Entry"""
    queryset = JournalEntry.objects.all()
    serializer_class = JournalEntryListSerializer
    permission_classes = [IsAuthenticated]
    filterset_fields = ['status']
    search_fields = ['entry_number', 'description', 'reference_number']
    ordering = ['-date', '-entry_number']
    pagination_class = None  # Non-paginated, frontend handles all data
    
    def get_serializer_class(self):
        if self.action == 'list':
            return JournalEntryListSerializer
        return JournalEntryDetailSerializer
    
    def get_queryset(self):
        queryset = JournalEntry.objects.select_related('created_by', 'approved_by').prefetch_related('items')
        
        # Filter by status
        status_filter = self.request.query_params.get('status')
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        
        # Filter by year
        year = self.request.query_params.get('year')
        if year:
            queryset = queryset.filter(date__year=year)
        
        # Filter by month
        month = self.request.query_params.get('month')
        if month:
            queryset = queryset.filter(date__month=month)
        
        return queryset
    
    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)
    
    @action(detail=True, methods=['post'])
    def post(self, request, pk=None):
        """Post journal entry"""
        journal = self.get_object()
        
        if journal.status != 'draft':
            return Response(
                {'error': 'Hanya draft yang bisa dipost!'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if journal.total_debit != journal.total_credit:
            return Response(
                {'error': 'Total debit harus sama dengan total credit!'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        journal.status = 'posted'
        journal.save()
        
        return Response({'message': 'Journal entry dipost!'})
    
    @action(detail=True, methods=['post'])
    def cancel(self, request, pk=None):
        """Cancel journal entry"""
        journal = self.get_object()
        
        if journal.status == 'canceled':
            return Response({'error': 'Sudah dicancel!'}, status=status.HTTP_400_BAD_REQUEST)
        
        journal.status = 'canceled'
        journal.save()
        
        return Response({'message': 'Journal entry dicancel!'})


class IncomeCategoryViewSet(viewsets.ModelViewSet):
    """ViewSet untuk Income Category"""
    queryset = IncomeCategory.objects.all()
    serializer_class = IncomeCategorySerializer
    permission_classes = [IsAuthenticated]
    filterset_fields = ['is_active']
    search_fields = ['name', 'code']


class IncomeViewSet(viewsets.ModelViewSet):
    """ViewSet untuk Income"""
    queryset = Income.objects.all()
    serializer_class = IncomeSerializer
    permission_classes = [IsAuthenticated]
    filterset_fields = ['category', 'status', 'customer', 'project']
    search_fields = ['income_number', 'description']
    ordering = ['-date']
    
    def get_queryset(self):
        queryset = Income.objects.select_related('category', 'account', 'customer', 'project', 'created_by')
        
        status_filter = self.request.query_params.get('status')
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        
        return queryset
    
    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)
    
    @action(detail=True, methods=['post'])
    def confirm(self, request, pk=None):
        """Confirm income"""
        income = self.get_object()
        
        if income.status != 'pending':
            return Response(
                {'error': 'Hanya pending yang bisa dikonfirm!'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        income.status = 'confirmed'
        income.save()
        
        return Response({'message': 'Income dikonfirm!'})
    
    @action(detail=True, methods=['post'])
    def complete(self, request, pk=None):
        """Complete income"""
        income = self.get_object()
        
        if income.status not in ['pending', 'confirmed']:
            return Response({'error': 'Status tidak valid!'}, status=status.HTTP_400_BAD_REQUEST)
        
        income.status = 'completed'
        income.is_completed = True
        income.save()
        
        return Response({'message': 'Income selesai!'})


class ExpenseCategoryViewSet(viewsets.ModelViewSet):
    """ViewSet untuk Expense Category"""
    queryset = ExpenseCategory.objects.all()
    serializer_class = ExpenseCategorySerializer
    permission_classes = [IsAuthenticated]
    filterset_fields = ['is_active']
    search_fields = ['name', 'code']


class ExpenseViewSet(viewsets.ModelViewSet):
    """ViewSet untuk Expense"""
    queryset = Expense.objects.all()
    serializer_class = ExpenseSerializer
    permission_classes = [IsAuthenticated]
    filterset_fields = ['category', 'status', 'vendor', 'project']
    search_fields = ['expense_number', 'description']
    ordering = ['-date']
    
    def get_queryset(self):
        queryset = Expense.objects.select_related('category', 'account', 'vendor', 'project', 'created_by')
        
        status_filter = self.request.query_params.get('status')
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        
        return queryset
    
    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)
    
    @action(detail=True, methods=['post'])
    def confirm(self, request, pk=None):
        """Confirm expense"""
        expense = self.get_object()
        
        if expense.status != 'pending':
            return Response(
                {'error': 'Hanya pending yang bisa dikonfirm!'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        expense.status = 'confirmed'
        expense.save()
        
        return Response({'message': 'Expense dikonfirm!'})
    
    @action(detail=True, methods=['post'])
    def complete(self, request, pk=None):
        """Complete expense"""
        expense = self.get_object()
        
        if expense.status not in ['pending', 'confirmed']:
            return Response({'error': 'Status tidak valid!'}, status=status.HTTP_400_BAD_REQUEST)
        
        expense.status = 'completed'
        expense.is_completed = True
        expense.save()
        
        return Response({'message': 'Expense selesai!'})


class InvoiceViewSet(viewsets.ModelViewSet):
    """ViewSet untuk Invoice"""
    queryset = Invoice.objects.all()
    permission_classes = [IsAuthenticated]
    filterset_fields = ['customer', 'project', 'status']
    search_fields = ['invoice_number']
    ordering = ['-date']
    
    def get_serializer_class(self):
        if self.action == 'list':
            return InvoiceListSerializer
        return InvoiceSerializer
    
    def get_queryset(self):
        queryset = Invoice.objects.select_related('customer', 'project', 'created_by')
        
        status_filter = self.request.query_params.get('status')
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        
        return queryset
    
    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)
    
    @action(detail=True, methods=['post'])
    def send(self, request, pk=None):
        """Kirim invoice"""
        invoice = self.get_object()
        
        if invoice.status != 'draft':
            return Response({'error': 'Status tidak valid!'}, status=status.HTTP_400_BAD_REQUEST)
        
        invoice.status = 'sent'
        invoice.save()
        
        return Response({'message': 'Invoice dikirim!'})


class PaymentViewSet(viewsets.ModelViewSet):
    """ViewSet untuk Payment"""
    queryset = Payment.objects.all()
    serializer_class = PaymentSerializer
    permission_classes = [IsAuthenticated]
    filterset_fields = ['customer', 'invoice', 'payment_type']
    search_fields = ['payment_number']
    ordering = ['-date']
    
    def get_queryset(self):
        queryset = Payment.objects.select_related('invoice', 'customer', 'account', 'created_by')
        
        customer = self.request.query_params.get('customer')
        if customer:
            queryset = queryset.filter(customer_id=customer)
        
        return queryset
    
    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)
        
        # Update invoice jika ada
        invoice = serializer.instance.invoice
        if invoice:
            invoice.amount_paid += serializer.instance.amount
            invoice.save()
    
    @action(detail=True, methods=['post'])
    def mark_paid(self, request, pk=None):
        """Tandai invoice lunas"""
        payment = self.get_object()
        
        if payment.invoice:
            payment.invoice.status = 'paid'
            payment.invoice.save()
        
        return Response({'message': 'Invoice lunas!'})


class BukuBesarViewSet(viewsets.ViewSet):
    """ViewSet untuk Buku Besar (General Ledger)"""
    permission_classes = [IsAuthenticated]
    
    def _get_queryset(self, request):
        """Get filtered journal entry items"""
        items = JournalEntryItem.objects.select_related(
            'journal_entry', 'account'
        ).filter(journal_entry__status='posted')
        
        # Filter by account
        account_id = request.query_params.get('account_id')
        if account_id:
            items = items.filter(account_id=account_id)
        
        # Filter by date range
        from_date = request.query_params.get('from_date')
        to_date = request.query_params.get('to_date')
        if from_date:
            items = items.filter(journal_entry__date__gte=from_date)
        if to_date:
            items = items.filter(journal_entry__date__lte=to_date)
        
        return items.order_by('account__code', 'journal_entry__date', 'journal_entry__id')
    
    def list(self, request):
        """
        Mendapatkan data Buku Besar (General Ledger).
        Data dikelompokkan per akun dengan detail transaksi.
        """
        items = self._get_queryset(request)
        
        # Kelompokkan per akun
        accounts_data = {}
        for item in items:
            acc_id = item.account_id
            if acc_id not in accounts_data:
                accounts_data[acc_id] = {
                    'account': {
                        'id': item.account.id,
                        'code': item.account.code,
                        'name': item.account.name,
                        'account_type': item.account.account_type,
                    },
                    'items': [],
                    'total_debit': 0,
                    'total_credit': 0,
                }
            accounts_data[acc_id]['items'].append({
                'id': item.id,
                'entry_number': item.journal_entry.entry_number,
                'date': item.journal_entry.date,
                'description': item.description or item.journal_entry.description,
                'debit': float(item.debit),
                'credit': float(item.credit),
            })
            accounts_data[acc_id]['total_debit'] += float(item.debit)
            accounts_data[acc_id]['total_credit'] += float(item.credit)
        
        # Hitung saldo awal dan saldo berjalan
        result = []
        for acc_id, data in accounts_data.items():
            running_balance = 0
            is_debit_normal = data['account']['account_type'] in ['asset', 'expense']
            
            for i, entry in enumerate(data['items']):
                if is_debit_normal:
                    running_balance += entry['debit'] - entry['credit']
                else:
                    running_balance += entry['credit'] - entry['debit']
                data['items'][i]['balance'] = round(running_balance, 2)
            
            data['saldo_akhir'] = round(running_balance, 2)
            data['total_debit'] = round(data['total_debit'], 2)
            data['total_credit'] = round(data['total_credit'], 2)
            result.append(data)
        
        # Urutkan berdasarkan kode akun
        result.sort(key=lambda x: x['account']['code'])
        
        return Response(result)
    
    @action(detail=False, methods=['get'])
    def accounts(self, request):
        """Dapatkan daftar akun untuk filter"""
        accounts = Account.objects.filter(is_active=True).order_by('code')
        return Response([{
            'id': a.id,
            'code': a.code,
            'name': a.name,
            'account_type': a.account_type,
        } for a in accounts])
    
    @action(detail=False, methods=['get'])
    def download(self, request):
        """Download Buku Besar dalam format Excel (.xlsx)"""
        file_format = request.query_params.get('format', 'xlsx')
        account_id = request.query_params.get('account_id')
        
        items = self._get_queryset(request)
        
        # Dapatkan nama akun untuk filename
        account_name = ''
        if account_id:
            try:
                acc = Account.objects.get(id=account_id)
                account_name = f" - {acc.code} {acc.name}"
            except Account.DoesNotExist:
                pass
        
        if file_format == 'csv':
            return self._download_csv(items, account_name)
        else:
            return self._download_xlsx(items, account_name)
    
    def _download_csv(self, items, account_name):
        """Download sebagai CSV"""
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Header
        writer.writerow(['No', 'Tanggal', 'No Jurnal', 'Keterangan', 'Debit', 'Kredit', 'Saldo'])
        
        # Data
        no = 1
        running_balance = 0
        prev_account = None
        total_debit = 0
        total_credit = 0
        
        for item in items:
            is_debit_normal = item.account.account_type in ['asset', 'expense']
            
            # Group header per akun
            if prev_account != item.account_id:
                if prev_account is not None:
                    writer.writerow([''])
                    writer.writerow(['', '', 'TOTAL', '', round(total_debit, 2), round(total_credit, 2), ''])
                    writer.writerow([''])
                
                writer.writerow([f'Akun: {item.account.code} - {item.account.name}'])
                writer.writerow(['No', 'Tanggal', 'No Jurnal', 'Keterangan', 'Debit', 'Kredit', 'Saldo'])
                
                running_balance = 0
                total_debit = 0
                total_credit = 0
                no = 1
            
            if is_debit_normal:
                running_balance += float(item.debit) - float(item.credit)
            else:
                running_balance += float(item.credit) - float(item.debit)
            
            writer.writerow([
                no,
                item.journal_entry.date.isoformat(),
                item.journal_entry.entry_number,
                item.description or item.journal_entry.description,
                round(float(item.debit), 2),
                round(float(item.credit), 2),
                round(running_balance, 2),
            ])
            
            total_debit += float(item.debit)
            total_credit += float(item.credit)
            no += 1
            prev_account = item.account_id
        
        # Last total
        if prev_account is not None:
            writer.writerow([''])
            writer.writerow(['', '', 'TOTAL', '', round(total_debit, 2), round(total_credit, 2), ''])
        
        response = HttpResponse(output.getvalue(), content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="buku_besar{account_name}.csv"'
        return response
    
    def _download_xlsx(self, items, account_name):
        """Download sebagai Excel (.xlsx)"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Buku Besar"
        
        # Style definitions
        header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
        header_fill = PatternFill(start_color='1B6EC2', end_color='1B6EC2', fill_type='solid')
        account_font = Font(name='Calibri', bold=True, size=12, color='1B6EC2')
        total_font = Font(name='Calibri', bold=True, size=11)
        total_fill = PatternFill(start_color='E8F0FE', end_color='E8F0FE', fill_type='solid')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )
        
        # Column widths
        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 14
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 40
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 18
        ws.column_dimensions['G'].width = 18
        
        # Title
        ws.merge_cells('A1:G1')
        title_cell = ws['A1']
        title_cell.value = 'BUKU BESAR - PT LENTERA ANUGERAH DIMENSI'
        title_cell.font = Font(name='Calibri', bold=True, size=14, color='1B6EC2')
        title_cell.alignment = Alignment(horizontal='center')
        ws.row_dimensions[1].height = 30
        
        row = 3
        no = 1
        running_balance = 0
        prev_account = None
        total_debit = 0
        total_credit = 0
        
        for item in items:
            is_debit_normal = item.account.account_type in ['asset', 'expense']
            
            # Group header per akun
            if prev_account != item.account_id:
                if prev_account is not None:
                    # Total for previous account
                    row += 1
                    ws.cell(row=row, column=2, value='').border = thin_border
                    ws.cell(row=row, column=3, value='TOTAL').font = total_font
                    ws.cell(row=row, column=3).fill = total_fill
                    ws.cell(row=row, column=3).border = thin_border
                    ws.cell(row=row, column=3).alignment = Alignment(horizontal='right')
                    ws.cell(row=row, column=5, value=total_debit).font = total_font
                    ws.cell(row=row, column=5).fill = total_fill
                    ws.cell(row=row, column=5).border = thin_border
                    ws.cell(row=row, column=5).number_format = '#,##0'
                    ws.cell(row=row, column=6, value=total_credit).font = total_font
                    ws.cell(row=row, column=6).fill = total_fill
                    ws.cell(row=row, column=6).border = thin_border
                    ws.cell(row=row, column=6).number_format = '#,##0'
                    row += 1
                
                row += 1
                # Account header
                ws.merge_cells(f'A{row}:G{row}')
                ws.cell(row=row, column=1, value=f"{item.account.code} - {item.account.name}").font = account_font
                row += 1
                
                # Column headers
                headers = ['No', 'Tanggal', 'No Jurnal', 'Keterangan', 'Debit', 'Kredit', 'Saldo']
                for col, h in enumerate(headers, 1):
                    cell = ws.cell(row=row, column=col, value=h)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center')
                    cell.border = thin_border
                
                running_balance = 0
                total_debit = 0
                total_credit = 0
                no = 1
            
            if is_debit_normal:
                running_balance += float(item.debit) - float(item.credit)
            else:
                running_balance += float(item.credit) - float(item.debit)
            
            row += 1
            ws.cell(row=row, column=1, value=no).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=1).border = thin_border
            ws.cell(row=row, column=2, value=item.journal_entry.date.isoformat()).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=2).border = thin_border
            ws.cell(row=row, column=3, value=item.journal_entry.entry_number).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=3).border = thin_border
            ws.cell(row=row, column=4, value=item.description or item.journal_entry.description).border = thin_border
            ws.cell(row=row, column=5, value=round(float(item.debit), 2))
            ws.cell(row=row, column=5).number_format = '#,##0'
            ws.cell(row=row, column=5).alignment = Alignment(horizontal='right')
            ws.cell(row=row, column=5).border = thin_border
            ws.cell(row=row, column=6, value=round(float(item.credit), 2))
            ws.cell(row=row, column=6).number_format = '#,##0'
            ws.cell(row=row, column=6).alignment = Alignment(horizontal='right')
            ws.cell(row=row, column=6).border = thin_border
            ws.cell(row=row, column=7, value=round(running_balance, 2))
            ws.cell(row=row, column=7).number_format = '#,##0'
            ws.cell(row=row, column=7).alignment = Alignment(horizontal='right')
            ws.cell(row=row, column=7).border = thin_border
            
            total_debit += float(item.debit)
            total_credit += float(item.credit)
            no += 1
            prev_account = item.account_id
        
        # Last total
        if prev_account is not None:
            row += 1
            ws.cell(row=row, column=3, value='TOTAL').font = total_font
            ws.cell(row=row, column=3).fill = total_fill
            ws.cell(row=row, column=3).border = thin_border
            ws.cell(row=row, column=3).alignment = Alignment(horizontal='right')
            ws.cell(row=row, column=5, value=round(total_debit, 2)).font = total_font
            ws.cell(row=row, column=5).fill = total_fill
            ws.cell(row=row, column=5).border = thin_border
            ws.cell(row=row, column=5).number_format = '#,##0'
            ws.cell(row=row, column=6, value=round(total_credit, 2)).font = total_font
            ws.cell(row=row, column=6).fill = total_fill
            ws.cell(row=row, column=6).border = thin_border
            ws.cell(row=row, column=6).number_format = '#,##0'
        
        # Save to response
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"buku_besar{account_name}.xlsx"
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


class FinanceReportViewSet(viewsets.ViewSet):
    """ViewSet untuk Laporan Keuangan"""
    permission_classes = [IsAuthenticated]
    
    @action(detail=False, methods=['get'])
    def summary(self, request):
        """Ringkasan keuangan"""
        # Total Income
        total_income = Income.objects.filter(
            is_completed=True
        ).aggregate(total=Sum('amount'))['total'] or 0
        
        # Total Expense
        total_expense = Expense.objects.filter(
            is_completed=True
        ).aggregate(total=Sum('amount'))['total'] or 0
        
        # Cash Balance
        cash_accounts = Account.objects.filter(is_cash=True)
        cash_balance = sum(account.balance for account in cash_accounts)
        
        # Piutang
        receivables = Invoice.objects.exclude(
            status__in=['paid', 'canceled']
        ).aggregate(total=Sum('amount_due'))['total'] or 0
        
        # Hutang
        payables = 0  # Calculate dari purchase orders jika ada
        
        return Response({
            'total_income': total_income,
            'total_expense': total_expense,
            'net_profit': total_income - total_expense,
            'cash_balance': cash_balance,
            'receivables': receivables,
            'payables': payables
        })
    
    @action(detail=False, methods=['get'])
    def profit_loss(self, request):
        """Laporan Rugi Laba"""
        from_date = request.query_params.get('from_date')
        to_date = request.query_params.get('to_date')
        
        incomes = Income.objects.filter(is_completed=True)
        expenses = Expense.objects.filter(is_completed=True)
        
        if from_date:
            incomes = incomes.filter(date__gte=from_date)
            expenses = expenses.filter(date__gte=from_date)
        
        if to_date:
            incomes = incomes.filter(date__lte=to_date)
            expenses = expenses.filter(date__lte=to_date)
        
        total_income = incomes.aggregate(total=Sum('amount'))['total'] or 0
        total_expense = expenses.aggregate(total=Sum('amount'))['total'] or 0
        
        return Response({
            'total_income': total_income,
            'total_expense': total_expense,
            'net_profit': total_income - total_expense,
            'profit_margin': ((total_income - total_expense) / total_income * 100) if total_income > 0 else 0
        })
    
    @action(detail=False, methods=['get'])
    def cash_flow(self, request):
        """Laporan Arus Kas"""
        from_date = request.query_params.get('from_date')
        to_date = request.query_params.get('to_date')
        
        cash_in = Income.objects.filter(is_completed=True)
        cash_out = Expense.objects.filter(is_completed=True)
        
        if from_date:
            cash_in = cash_in.filter(date__gte=from_date)
            cash_out = cash_out.filter(date__gte=from_date)
        
        if to_date:
            cash_in = cash_in.filter(date__lte=to_date)
            cash_out = cash_out.filter(date__lte=to_date)
        
        return Response({
            'cash_in': cash_in.aggregate(total=Sum('amount'))['total'] or 0,
            'cash_out': cash_out.aggregate(total=Sum('amount'))['total'] or 0,
            'net_cash_flow': (
                cash_in.aggregate(total=Sum('amount'))['total'] or 0
            ) - (
                cash_out.aggregate(total=Sum('amount'))['total'] or 0
            )
        })
